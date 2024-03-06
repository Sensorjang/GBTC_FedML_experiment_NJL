import collections
import copy
import logging
import random
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor

from scipy.optimize import minimize, Bounds
from scipy.stats import lognorm
import matplotlib.pyplot as plt
import numpy as np
import torch
import wandb
from fedml import mlops
from fedml.ml.trainer.trainer_creator import create_model_trainer
from torch.utils.data import DataLoader

from .client import Client
import openpyxl
import time

# 常超参数
interval_params = {
    'miu': (15, 20),  # 处理单位样本所需的CPU周期数
    'p': (0.5, 1),  # 客户端发射功率
    'f': (0.05, 0.4),  # 客户端的cpu频率范围
    'Z': -114,  # 噪声Z0
    'B': 10,  # 总上行带宽B
    'delta': 6272,  # 单位数据样本大小（每个样本的物理大小/bit）
    'M': 3.4 * 10 ** 6,  # 模型大小
    'sigma': 1,  # 小尺度衰落，瑞利分布的方差
    'rho': 8,  # 阴影衰落， 对数正态分布的标准差
    'e': 1e-28,  # 计算芯片组有效电容参数
    'd': (0, 0.08),  # 客户端与联邦服务器的物理距离/km
    'k': 30,  # 初始联盟的个数
    'omega': 900,  # 联盟主对支付成本与训练时间的权衡系数,
    'lambda_cm': (1e-8, 1e-7),  # 客户单位通信开销范围
    'lambda_cp': (1e-8, 1e-7),  # 客户单位计算开销范围
    'trust_range': (0, 10),  # 信任值范围
    'distrust_probability': 0.01,  # 完全不信任的概率
    'trust_threshold': 35,  # 联盟信任阈值（低于阈值的剔除）
    'member_tolerance': 3,  # 联盟成员数量的容忍度（低于阈值的剔除）
    'pay_range': (10, 100),  # 盟主支付元素的范围
    'data_imp': (0.6, 0.9),  # 盟主对数据质量的重视程度
    'data_share': (0.05, 0.1),  # 成员对盟主的数据共享率
    # 'N': (20, 30),  # 客户端数量范围
    # 'D': (200, 1600),  # 数据集大小范围
    # 'I': 5,  # 本地训练epoch大小
    # 'batch_size': 10,  # 本地训练批大小
    # 'lr': 0.01,  # 本地训练学习率
}  # 注释掉的全部到yaml中设置
quality_weights = {'cpu': 0.05, 'ram': 0.05, 'bm': 0.05, 'q': 0.85}  # 重视数据质量
quality_ranges = {'cpu': (1, 5), 'ram': (1, 5), 'bm': (1, 5)}
accuracy_list = []
loss_list = []
train_time_list = []
union_num_list = []
member_num_list = []
selfish_num_list = []

# 参数0
# 统计global_client_num_in_total个客户每个人的被选择次数
plt.figure(1, figsize=(16, 4))

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 删除默认创建的工作表（如果需要的话）
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])
# 创建一个工作表用于存放全局指标
global_metrics_ws = wb.create_sheet('Global Metrics')
# 写入全局指标的标头行
global_metrics_ws.append(['Round', 'Test Accuracy', 'Test Loss', 'Training Time', 'Union Num', 'Member Num', 'Selfish Num'])

interval = 5


class AutoIncrementDict:
    def __init__(self):
        self._data = {}
        self._next_id = 0

    def add(self, value):
        self._data[self._next_id] = value
        current_id = self._next_id
        self._next_id += 1
        return current_id

    def remove(self, key):
        if key in self._data:
            del self._data[key]

    def get(self, key):
        return self._data.get(key, None)

    def __getitem__(self, key):
        return self.get(key)

    def __delitem__(self, key):
        self.remove(key)

    def __repr__(self):
        return repr(self._data)

    def items(self):
        return self._data.items()

    def keys(self):
        return self._data.keys()

    def values(self):
        return self._data.values()

    def remove_empty_values(self):
        keys_to_remove = [key for key, value in self._data.items() if not value]
        for key in keys_to_remove:
            del self._data[key]

    def ban_union(self, keys_to_remove):
        for key in keys_to_remove:
            del self._data[key]


class FedGBTCAPI(object):  # 变量参考FARA代码
    def __init__(self, args, device, dataset, model):
        self.device = device
        self.args = args
        [
            train_data_num,
            test_data_num,
            train_data_global,
            test_data_global,
            train_data_local_num_dict,
            train_data_local_dict,
            test_data_local_dict,
            class_num,
        ] = dataset

        self.train_global = train_data_global
        self.test_global = test_data_global
        self.val_global = None
        self.train_data_num_in_total = train_data_num
        self.test_data_num_in_total = test_data_num
        self.class_num = class_num
        self.client_num_in_total = 50 # 不要太多了，否则求解时间太长
        self.client_list = []  # 原始的客户集合（全部），不以联盟区分
        self.train_data_local_num_dict = self.get_local_sample_num(train_data_local_dict)
        self.train_data_local_dict = train_data_local_dict
        self.test_data_local_dict = test_data_local_dict
        # 乐享联盟参数
        self.common_params = {}
        self.quality_weights = quality_weights
        self.client_data_distribution = {i: {} for i in range(self.client_num_in_total)}  # 存储每个客户的数据概率分布
        self.client_params = {i: {} for i in range(self.client_num_in_total)}  # 存储每个客户的交互参数
        self.client_quality_list = {  # 存储客户的特征向量（质量属性）
            i: {'cpu': np.random.randint(low=quality_ranges['cpu'][0], high=quality_ranges['cpu'][1] + 1),
                'ram': np.random.randint(low=quality_ranges['ram'][0], high=quality_ranges['ram'][1] + 1),
                # 其余打酱油的属性随机生成，数据质量后面会计算
                'bm': np.random.randint(low=quality_ranges['bm'][0], high=quality_ranges['bm'][1] + 1), 'q': 0.0}
            for i in range(self.client_num_in_total)}
        self.trust_matrix = None  # 初始化直接生成信任矩阵，全局静态
        self.client_unions = {}  # 稳定的联盟结构（可能包含多个联盟，键为联盟主的id，值为联盟成员的id集合，暂不考虑联盟结构的变化）
        self.time_unions = {}  # 记录每轮每个稳定联盟的整体训练时间，以盟主id为键
        self.imp_unions = {}  # 记录每个联盟对不同资源的重视程度(元组字典)
        # 历史联盟(客户历史所参与的,还没有盟主，字典存放联盟的id-客户对其偏好值)
        self.his_client_unions = {i: {} for i in range(self.client_num_in_total)}
        self.client_banned_list = {}  # 存储被剔除的客户（键为id、值为 1. 不成盟 2. 累计信任值不足 3. 均衡解为0）
        self.client_rewards = {}  # 存储每个客户的奖励值

        logging.info("model = {}".format(model))
        self.model_trainer = create_model_trainer(model, args)
        self.model = model
        logging.info("self.model_trainer = {}".format(self.model_trainer))

        self._setup_clients(train_data_local_dict, test_data_local_dict)

    def get_local_sample_num(self, train_data_local_dict):  # 现在传入的只会是dataloader对象
        train_data_local_num_dict = {}
        for cid in range(self.client_num_in_total):
            train_data = train_data_local_dict[cid]
            sample_num = len(train_data.dataset)
            train_data_local_num_dict[cid] = sample_num
        return train_data_local_num_dict

    def params_generator(self):  # 客户独立参数的赋值
        for key, value in interval_params.items():
            if key == 'sigma':
                rl = np.random.rayleigh(scale=value, size=self.client_num_in_total)
                for i, sigma_i in enumerate(rl):
                    self.client_params[i]['sigma'] = sigma_i
            elif key == 'rho':
                ln = lognorm(s=value)
                for i, rho_i in enumerate(ln.rvs(size=self.client_num_in_total)):
                    self.client_params[i]['rho'] = rho_i
            elif key == 'pay_range':  # 支付元素的大小范围，转存到self
                self.common_params['pay_range'] = value
            elif key == 'data_imp':  # 数据质量重视程度范围，转存到self
                self.common_params['data_imp'] = value
            elif key == 'data_share':  # 数据共享率范围，转存到self
                self.common_params['data_share'] = value
            elif key == 'trust_range':  # 信任值范围，转存到self
                self.common_params['trust_range'] = value
            elif isinstance(value, tuple):
                for i in range(self.client_num_in_total):
                    self.client_params[i][key] = np.random.uniform(*value)
            else:
                self.common_params[key] = value

        for client in self.client_list:  # 客户本地数据量的赋值
            self.client_params[client.client_idx]['D'] = client.get_sample_number()
        self.common_params['I'] = self.args.epochs

    def create_trust_graph(self):
        """
        创建信任图。初始化时除对角线外的元素在指定范围内随机生成。
        之后以一定概率将部分值设置为负无穷，表示完全不信任，并且设置某些元素为0，表示不认识（避免掩码冲突）。
        :param value_range: 信任值的范围，为(min_value, max_value)
        :param distrust_probability: 完全不信任的概率
        :param unknown_probability: 不认识的概率
        :return: 信任图，包含不信任和不认识关系的随机值，对角线元素为0
        """
        distrust_probability = self.common_params['distrust_probability']
        value_range = self.common_params['trust_range']
        shape = (self.client_num_in_total, self.client_num_in_total)
        min_value, max_value = value_range
        # 初始化信任矩阵，除对角线外的元素在指定范围内随机生成
        trust_array = np.random.uniform(min_value, max_value, size=shape)
        # 将对角线元素设置为0
        np.fill_diagonal(trust_array, 0)
        # 生成不信任掩码，并应用，设置部分元素为负无穷
        distrust_mask = np.random.rand(*shape) < distrust_probability
        np.fill_diagonal(distrust_mask, False)  # 保持对角线上的元素为0
        trust_array[distrust_mask] = -np.inf

        return trust_array

    def upgrade_union_uid(self, union, uid_old, uid_new, nid=None):  # 新旧联盟的交替(通常发生在某客户离开/加入某联盟，会产生新的联盟，此时需要修改其余成员的记录)
        for cid in union:  # 这里可选则将新增客户id传入,避免删除旧联盟id记录的操作
            if cid != nid:
                self.his_client_unions[cid].pop(uid_old)
            self.his_client_unions[cid][uid_new] = self.cal_preference(cid, union)

    def unions_formation(self):
        """
        生成初始联盟划分，同时避免联盟中存在信任值为负无穷的客户关系。
        """
        customer_ids = np.arange(self.client_num_in_total)
        np.random.shuffle(customer_ids)
        unions = AutoIncrementDict()
        # 初始化联盟字典（自动计数id）
        for customer_id in customer_ids:
            added = False
            for uid, members in unions.items():
                can_add = True
                for member in members:
                    if (self.trust_matrix[customer_id, member] == -np.inf or
                            self.trust_matrix[member, customer_id] == -np.inf):
                        can_add = False
                        break
                if can_add:
                    unions[uid].append(customer_id)
                    self.his_client_unions[customer_id][uid] = None  # 此时联盟还未稳定，先不计算偏好值
                    added = True
                    break
            if not added:
                new_uid = unions.add([customer_id])
                self.his_client_unions[customer_id][new_uid] = None  # 此时联盟还未稳定，先不计算偏好值
        unions.remove_empty_values()  # 清除空联盟

        stable = False
        while not stable:
            stable = True  # 假设本轮次不再发生联盟变化
            for cid in range(self.client_num_in_total):  # 遍历每一位客户,初始将目前定义为最优
                uid_i = next(reversed(self.his_client_unions[cid].keys()))  # 找到当前客户i所在的联盟id
                best_pre_i = self.cal_preference(cid, unions[uid_i])  # 先计算客户对当前联盟的偏好值
                best_uid_i = uid_i
                for uid, union in unions.items():
                    if uid == best_uid_i:  # 避免重新评估当前联盟
                        continue
                    else:
                        pre = self.cal_preference(cid, union) \
                            if uid not in self.his_client_unions[cid] else 0  # 计算客户对该新联盟的偏好值（此时需要考虑历史联盟的问题）
                        if pre > best_pre_i:
                            best_pre_i = pre
                            best_uid_i = uid
                            stable = False  # 如果发生了联盟变化，则本轮次不再稳定

                if best_uid_i != uid_i:  # 如果找到了更好的联盟,双更新,也要更新原来两个联盟中客户绑定的uid
                    # 更新旧联盟
                    union_former = unions[uid_i]
                    union_former.remove(cid)
                    unions.remove(uid_i)
                    new_uid_former = unions.add(union_former)
                    self.upgrade_union_uid(union_former, uid_i, new_uid_former)
                    # 更新新联盟
                    union_latter = unions[best_uid_i]
                    unions.remove(best_uid_i)
                    union_latter.append(cid)
                    new_uid_latter = unions.add(union_latter)
                    self.his_client_unions[cid].pop(uid_i)
                    self.upgrade_union_uid(union_latter, best_uid_i, new_uid_latter, cid)
                unions.remove_empty_values()  # 清除空联盟

        # 新增代码：在最终联盟确定后，清除只有一个客户的联盟
        for uid, union in unions.items():  # 使用list来避免在遍历时修改字典
            ban_cid_list = []
            for cid in union:  # 清除社会信任不足阈值的成员
                if self.cal_trust_sum(union, cid) < self.common_params['trust_threshold']:
                    self.his_client_unions[cid].pop(uid)
                    self.client_banned_list[cid] = 1  # 因为社会信任不足被剔除
                    ban_cid_list.append(cid)
            for cid in ban_cid_list:
                union.remove(cid)

        # 清除自私集群的联盟
        ban_union_key_list = []
        for uid, union in unions.items():
            if len(union) <= self.common_params['member_tolerance']:
                for cid in union:
                    self.his_client_unions[cid].pop(uid)
                    self.client_banned_list[cid] = 2
                ban_union_key_list.append(uid)
        print(ban_union_key_list)
        unions.ban_union(ban_union_key_list)

        # 返回稳定的联盟
        return unions

    def cal_client_quality(self, cid):  # 计算客户i的加权质量属性值
        score = 0.0
        for key, value in self.client_quality_list[cid].items():
            score += value * self.quality_weights[key]
        return score

    def cm_election(self, unions):  # 联盟主选举算法
        for uid, union in unions.items():
            best_score = -np.inf
            best_cid = -1
            for cid in union:
                score = self.cal_client_quality(cid)
                if score > best_score:
                    best_score = score
                    best_cid = cid
            if best_cid == -1:
                raise ValueError("No client in the union")
            # 从联盟中移除联盟主，准备更新联盟成员信息
            self.client_unions[best_cid] = [cid for cid in union if cid != best_cid]
            self.time_unions[best_cid] = {}  # 初始化每个稳定联盟的时间记录容器，以round_idx为键
            data_imp = np.random.uniform(*self.common_params['data_imp'])
            self.imp_unions[best_cid] = (1 - data_imp, data_imp)
            self.cm_data_share(best_cid)  # 联盟内部数据贡献

    def cm_data_share(self, u_cid):  # 联盟内部数据贡献
        ally_clinet = self.client_list[u_cid]
        for cid in self.client_unions[u_cid]:
            client = self.client_list[cid]
            share_rate = np.random.uniform(*self.common_params['data_share'])
            share_data, share_number = client.share_data_up(share_rate)
            ally_clinet.get_shared_data(cid, share_data, share_number)
        # 盟主装载联盟共享数据集
        ally_clinet.update_share_data()

    def cal_preference(self, i, union):  # 客户对联盟的偏好函数,暂时不考虑历史联盟 (计算时不考虑历史参与)
        pre = 0.0
        for j in union:
            if self.trust_matrix[i][j] == -np.inf:
                pre = -np.inf
                break
            # elif
            else:
                pre += self.trust_matrix[i][j]
        return pre

    def cal_trust_sum(self, union, cid):
        return np.sum([self.trust_matrix[i][cid] for i in union])

    def cal_data_quality(self, u_cid):  # 计算联盟内成员的数据质量(这的数据质量不太一样，是按类别在整体的权熵，不是EMD) 所需的参数全部由客户提供
        entropy = {}  # 暂存每位客户数据概率分布的熵
        label_num_per_client = {}  # 统计联盟内每个客户的每个类别的样本量
        for cid in self.client_unions[u_cid]:
            label_num_per_client[cid] = []
            total_sample_num = self.client_list[cid].get_sample_number()
            for j in range(self.class_num):
                this_sample_num = self.client_list[cid].get_sample_class_number(j)
                label_num_per_client[cid].append((this_sample_num + 1)
                                                 / total_sample_num + self.class_num)
        # 计算每位客户的类别熵权
        for cid in self.client_unions[u_cid]:
            sum_entropy = 0
            for j in range(self.class_num):
                sum_entropy -= label_num_per_client[cid][j] * np.log(label_num_per_client[cid][j])
            entropy[cid] = sum_entropy
        sum_entropy = sum([(1 - entropy) for entropy in list(entropy.values())])
        wi = {cid: (1 - entropy) / sum_entropy for cid, entropy in entropy.items()}
        # 计算并记录每位客户的数据质量
        for cid, eweight in wi.items():
            self.client_quality_list[cid]['q'] = eweight * sum(label_num_per_client[cid])

    def cal_client_utility(self, u_cid, bandwidth_vector, payment_vector):  # 计算联盟内所有客户的效用(传入初始分配默认以顺序作为成员索引)
        union = self.client_unions[u_cid]
        print(union)
        utilities = {}
        for i, cid in enumerate(union):
            pay = payment_vector[i]
            band_per = bandwidth_vector[i]
            c_cm = self.client_params[cid]['lambda_cm'] * self._cal_cm_time_flag(cid) / band_per
            c_cp = (self.client_params[cid]['lambda_cp']
                    * self._cal_cp_time_flag(cid) / self.client_params[cid]['f']
                    * self.common_params['e']
                    * self.client_params[cid]['f'] ** 3)
            utilities[cid] = (self.imp_unions[u_cid][0] * self.client_params[cid]['f'] + self.imp_unions[u_cid][1]
                              * self.client_quality_list[cid]['q'] * pay - c_cm - c_cp)
        return utilities

    def ms_obj(self, x, uid, flags):  # 优化变量分别为：支付向量、带宽分配向量，整合到一个决策变量中(便于优化器)
        n = len(x) // 2
        price = x[:n]  # 解耦出两类决策变量
        band = x[n:2 * n]
        # print(band)
        t_cm = []
        t_cp = copy.deepcopy(flags['time_cp'])
        data_values = flags['data_value']
        cp_values = flags['cp_value']
        values = []
        for i, cid in enumerate(self.client_unions[uid]):
            t_cm.append(flags['time_cm'] / band[i])
            values.append((cp_values[i] * price[i] + data_values[i]) * price[i])
        t = np.array(t_cm) + np.array(t_cp)
        t_max = np.max(t)
        obj = np.sum(values) + t_max * self.common_params['omega']
        return obj

    def _calculate_real_bound(self, uid, cid):
        flag = (2 * self.client_params[cid]['lambda_cp'] * self.client_params[cid]['miu']
                * self.common_params['delta'] * self.client_params[cid]['D']
                * self.common_params['I'] / self.imp_unions[uid][0])
        return flag

    def _cal_cp_value_flag(self, pre, cid):
        return (pre ** 2 / (2 * self.client_params[cid]['lambda_cp']
                            * self.client_params[cid]['miu'] * self.common_params['delta']
                            * self.client_params[cid]['D'] * self.common_params['I'] * self.common_params['e']))

    def _cal_data_value_flag(self, pre, cid):
        return self.client_quality_list[cid]['q'] * pre

    def _cal_cm_time_flag(self, cid):
        g_i = self.client_params[cid]['rho'] * self.client_params[cid]['sigma'] * (
                128.1 + 37.6 * np.log2(self.client_params[cid]['d']))
        return (self.common_params['M'] /
                (self.common_params['B'] * np.log2(1 + g_i * self.client_params[cid]['p']
                                                   / self.common_params['Z'])))

    def _cal_cp_time_flag(self, cid):  # 2024-3-3 为适应两种计算(f不同)，现去掉f变量
        return (self.client_params[cid]['miu'] * self.common_params['delta'] * self.client_params[cid]['D']
                * self.common_params['I'])

    def ms_game_solution(self, u_cid):  # 主从博弈最优求解（按盟主id来）不用计算时间开销，问题定义的比较清晰，只用带入
        # 先生成一组随机分配(支付向量、带宽分配比向量)
        pay_vector, band_vector = [], []
        member_count = 0
        pay_min = self.common_params['pay_range'][0]
        pay_max = self.common_params['pay_range'][1]
        f_max = interval_params['f'][1]
        f_min = interval_params['f'][0]
        ubound_pay, lbound_pay = [], []
        for cid in self.client_unions[u_cid]:  # 先确定每位客户的支付元素的范围，随机生成的同时保存这个范围便于约束条件的设置
            band_vector.append(np.random.rand())
            flag = self._calculate_real_bound(u_cid, cid)
            pay_item_min = min(pay_min, f_min * flag)
            pay_item_max = max(pay_max, f_max * flag)
            pay_vector.append(np.random.uniform(pay_item_min, pay_item_max))
            print((pay_item_max, pay_item_min))
            ubound_pay.append(pay_item_max)
            lbound_pay.append(pay_item_min)
            member_count += 1
        band_vector = np.array(pay_vector)
        band_vector = band_vector / np.sum(band_vector)  # 约束范围转np，便于》。。
        lbound_pay, ubound_pay = np.array(lbound_pay), np.array(ubound_pay)

        # 第二阶段：求出初始分配下，成员的均衡解-计算每位成员此时的效用，根据效用取值决定其均衡解
        # 先计算联盟中每个成员的数据质量：
        self.cal_data_quality(u_cid)
        # 然后计算每位客户此时的效用
        utility_per_client = self.cal_client_utility(u_cid, band_vector, pay_vector)
        strategy_per_client = {}  # 每位成员最终的付出值(二阶段的均衡解)
        cmp_flags, data_flags = [], []  # 记录两阶段重复计算（目标函数中，计算价值、数据价值）
        time_cm_flags, time_cp_flags = [], []  # （约束条件中，通信/计算时间）
        for i, (cid, u) in enumerate(utility_per_client.items()):  # 根据当前效用值来确定每位成员的均衡解
            cmp_flags.append(self._cal_cp_value_flag(self.imp_unions[u_cid][0], cid))
            data_flags.append(self._cal_data_value_flag(self.imp_unions[u_cid][1], cid))
            time_cm_flags.append(self._cal_cm_time_flag(cid))
            f_flag = cmp_flags[-1] * pay_vector[i]  # 2024-3-3 需求有变，去掉0决策
            if f_flag >= f_max:
                strategy_per_client[cid] = f_max
            elif f_flag <= f_min:
                strategy_per_client[cid] = f_min
            else:
                strategy_per_client[cid] = f_flag  # 并且决策变量f*代替一阶段优化问题的f
            time_cp_flags.append(self._cal_cp_time_flag(cid) / strategy_per_client[cid])
        # 第一阶段，求出当前盟主的均衡解, 调用minimize函数时传入额外参数
        x0 = np.concatenate((pay_vector, band_vector))
        flags = {'cp_value': cmp_flags, 'data_value': data_flags, 'time_cm': time_cm_flags, 'time_cp': time_cp_flags}
        # 为带宽变量创建边界（注意：这里用了稍微大于0的下界和小于1的上界来确保不触及边界）
        lbound_bandwidth = [1e-8] * member_count  # 带宽变量的下界
        ubound_bandwidth = [1 - 1e-8] * member_count  # 带宽变量的上界
        # 合并边界
        lbound = np.concatenate((lbound_pay, lbound_bandwidth))
        ubound = np.concatenate((ubound_pay, ubound_bandwidth))
        bounds = Bounds(lbound, ubound)
        solution = minimize(
            fun=self.ms_obj,
            x0=x0,
            args=(u_cid, flags),
            method='trust-constr',  # 这里为什么报类型错误不是很清楚，官方文档里写的可以字典或者字典列表
            bounds=bounds,
            constraints={'type': 'eq', 'fun': lambda x: sum(x[member_count:2 * member_count]) - 1},
            tol=1e-3,
            options={'disp': True}
        )
        # 检查解决方案
        if solution.success:
            print("Optimal solution for union of {} found.".format(u_cid))
            # 最优解
            optimal_x = solution.x
            pay_vector = optimal_x[:member_count]
            band_vector = optimal_x[member_count:2 * member_count]
            # 目标函数的最优值
            optimal_fun = solution.fun
        else:
            print("Optimal solution for union of {} failed.".format(u_cid))
        # print(strategy_per_client)  # 2024-3-3 由于已经没有0决策，以下弃用
        # for i, cid in enumerate(self.client_unions[u_cid]):  # 每位成员最终的分配值(一阶段的均衡解)
        #     if strategy_per_client[cid] > 0:
        #         self.client_rewards[cid] = (pay_vector[i], band_vector[i])
        #     else:
        #         self.client_rewards[cid] = (0, 0)  # 联盟内部均衡解不淘汰
                # self.client_banned_List[cid] = 3  # 因为均衡解为0被剔除(弃用)

    def _setup_clients(self, train_data_local_dict, test_data_local_dict):
        logging.info("############setup_clients (START)#############")
        # 参数1
        for client_idx in range(self.client_num_in_total):
            c = Client(
                client_idx,
                train_data_local_dict[client_idx],
                test_data_local_dict[client_idx],
                self.train_data_local_num_dict[client_idx],
                self.args,
                self.device,
                copy.deepcopy(self.model_trainer),  # 深复制一下比较好
            )
            self.client_list.append(c)
        # 客户交互属性的生成
        self.params_generator()
        logging.info("############setup_clients (END)#############")


    def train_union(self, u_cid, w_global):
        union = self.client_unions[u_cid]
        num_union = 0  # 记录联盟内的总样本量
        w_locals = []  # 暂存联盟内客户端返回的模型
        for cid in union + [u_cid]:
            client = self.client_list[cid]
            mlops.event("train", event_started=True, event_edge_id=u_cid)
            w = client.train(copy.deepcopy(w_global))
            mlops.event("train", event_started=True, event_edge_id=u_cid)
            num_i = self.train_data_local_num_dict[cid]
            num_union += num_i
            w_locals.append((num_i, copy.deepcopy(w)))
        return num_union, w_locals


    def train(self):
        logging.info("self.model_trainer = {}".format(self.model_trainer))
        w_global = self.model_trainer.get_model_params()
        mlops.log_training_status(mlops.ClientConstants.MSG_MLOPS_CLIENT_STATUS_TRAINING)
        mlops.log_aggregation_status(mlops.ServerConstants.MSG_MLOPS_SERVER_STATUS_RUNNING)
        mlops.log_round_info(self.args.comm_round, -1)
        # 乐享联盟
        mlops.event("信任图生成", event_started=True)
        self.trust_matrix = self.create_trust_graph()  # 客户交互图生成
        print(self.trust_matrix)
        mlops.event("信任图生成", event_started=False)
        mlops.event("联盟生成", event_started=True)
        unions = self.unions_formation()  # 初始联盟划分，并得到最优化的划分
        print(unions)
        mlops.event("联盟生成", event_started=False)
        mlops.event("联盟主选举", event_started=True)
        self.cm_election(unions)  # 联盟主选举，并得到完整的联盟划分
        print(self.client_unions)
        mlops.event("联盟主选举", event_started=False)
        print(len(self.client_banned_list))
        # 主从博弈
        for u_cid, union in self.client_unions.items():  # 先把两阶段的东西全部弄完，主要是淘汰掉失信客户
            mlops.event("主从博弈求解", event_started=True, event_value="盟主：{}".format(str(u_cid)))
            self.ms_game_solution(u_cid)  # 完成主从博弈最优求解（求出带宽、支付分配）
            mlops.event("主从博弈求解", event_started=False, event_value="盟主：{}".format(str(u_cid)))
        print(len(self.client_banned_list))

        for round_idx in range(self.args.comm_round):
            logging.info("################Communication round : {}".format(round_idx))
            w_unions = {}  # 暂存联盟主返回的模型，用于全局聚合， 以联盟主id为键
            time_s = time.time()  # 记录训练开始时间, 多线程并行运行每个联盟的训练
            with ThreadPoolExecutor(max_workers=10) as executor:
                futures = {}
                for u_cid in self.client_unions:
                    # 提交任务到线程池
                    future = executor.submit(self.train_union, u_cid, w_global)
                    futures[u_cid] = future
                # 等待所有任务完成
                for u_cid, future in futures.items():  # 存入：(联盟总样本量、w_locals)
                    u_sample_num, w_locals = future.result()
                    w_unions[u_cid] = (u_sample_num, w_locals)
            time_e = time.time()  # 记录训练开始时间, 多线程并行运行每个联盟的训练
            train_time_list.append(time_e - time_s)

            # 联盟聚合
            w_final = []
            for u_cid, (u_sample_num, w_locals) in w_unions.items():
                mlops.event("Agg", event_started=True, event_value="轮次{}_联盟主{}".format(str(round_idx), str(u_cid)))
                w_final.append((u_sample_num, self._aggregate(w_locals)))
                mlops.event("Agg", event_started=True, event_value="轮次{}_联盟主{}".format(str(round_idx), str(u_cid)))

            # 全局聚合
            mlops.event("Agg", event_started=True, event_value="轮次{}_全局".format(str(round_idx)))
            w_global = self._aggregate(w_final)
            self.model_trainer.set_model_params(w_global)
            mlops.event("Agg", event_started=False, event_value="轮次{}_全局".format(str(round_idx)))

            # test results
            # at last round
            train_acc, train_loss, test_acc, test_loss = 0, 0, 0, 0
            if round_idx == self.args.comm_round - 1:
                self._local_test_on_all_clients(round_idx)
            # per {frequency_of_the_test} round
            elif round_idx % self.args.frequency_of_the_test == 0:
                if self.args.dataset.startswith("stackoverflow"):
                    self._local_test_on_validation_set(round_idx)
                else:
                    self._local_test_on_all_clients(round_idx)
            # 记录联盟实时数据
            union_num_list.append(len(self.client_unions))
            selfish_num = len(self.client_banned_list)
            selfish_num_list.append(selfish_num)
            member_num_list.append(self.client_num_in_total - selfish_num)

            mlops.log_round_info(self.args.comm_round, round_idx)
            # 休眠一段时间，以便下一个循环开始前有一些时间
            # time.sleep(interval)

        # 填充数据到工作表
        for i in range(len(accuracy_list)):
            # 轮次号，测试精度，测试损失，训练时间
            round_num = i + 1  # 轮次号从1开始
            accuracy = accuracy_list[i]
            loss = loss_list[i]
            train_time = train_time_list[i]
            union_num = union_num_list[i]
            member_num = member_num_list[i]
            selfish_num = selfish_num_list[i]
            global_metrics_ws.append([round_num, accuracy, loss, train_time, union_num, member_num, selfish_num])

        # 保存Excel文件到self.args.excel_save_path+文件名
        wb.save(
            self.args.excel_save_path + self.args.model + "_[" + self.args.dataset + "]_GBTC_training_results_NIID" + str(
                self.args.experiment_niid_level) + ".xlsx")

        mlops.log_training_finished_status()
        mlops.log_aggregation_finished_status()

    def _generate_validation_set(self, num_samples=10000):
        test_data_num = len(self.test_global.dataset)
        sample_indices = random.sample(range(test_data_num), min(num_samples, test_data_num))
        subset = torch.utils.data.Subset(self.test_global.dataset, sample_indices)
        sample_testset = torch.utils.data.DataLoader(subset, batch_size=self.args.batch_size)
        self.val_global = sample_testset

    def _aggregate(self, w_locals):
        training_num = 0
        for idx in range(len(w_locals)):
            (sample_num, averaged_params) = w_locals[idx]
            training_num += sample_num

        (sample_num, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            print("______k = " + str(k))
            for i in range(0, len(w_locals)):
                local_sample_number, local_model_params = w_locals[i]
                w = local_sample_number / training_num
                if i == 0:
                    averaged_params[k] = local_model_params[k] * w
                else:
                    averaged_params[k] += local_model_params[k] * w
        return averaged_params

    def _aggregate_resnet(self, w_locals, eweights):  # 弃用
        eweights_sum = sum(eweights)

        averaged_params = OrderedDict()
        for (_, local_model_params), w in zip(w_locals, eweights):
            for key, param in local_model_params.items():
                if key not in averaged_params:
                    averaged_params[key] = []

                    # 获取对应权重
                local_wight = w / eweights_sum

                # 根据类型分别计算加权平均
                if 'conv' in key:
                    averaged_params[key] += local_wight * param.clone().detach()
                elif 'bn' in key:
                    averaged_params[key] += local_wight * param.clone().detach()
                elif 'fc' in key:
                    averaged_params[key] += local_wight * param.clone().detach()

        for key in averaged_params:
            averaged_params[key] /= eweights_sum

        return averaged_params

    def _aggregate_noniid_avg(self, w_locals):
        """
        The old aggregate method will impact the model performance when it comes to Non-IID setting
        Args:
            w_locals:
        Returns:
        """
        (_, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            temp_w = []
            for (_, local_w) in w_locals:
                temp_w.append(local_w[k])
            averaged_params[k] = sum(temp_w) / len(temp_w)
        return averaged_params

    def _local_test_on_all_clients(self, round_idx):

        logging.info("################local_test_on_all_clients : {}".format(round_idx))

        train_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        test_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        client = self.client_list[0]

        for client_idx in range(self.args.client_num_in_total):
            """
            Note: for datasets like "fed_CIFAR100" and "fed_shakespheare",
            the training client number is larger than the testing client number
            """
            if self.test_data_local_dict[client_idx] is None:
                continue
            client.update_local_dataset(
                self.train_data_local_dict[client_idx],
                self.test_data_local_dict[client_idx],
                self.train_data_local_num_dict[client_idx],
            )
            # train data
            train_local_metrics = client.local_test(False)
            train_metrics["num_samples"].append(copy.deepcopy(train_local_metrics["test_total"]))
            train_metrics["num_correct"].append(copy.deepcopy(train_local_metrics["test_correct"]))
            train_metrics["losses"].append(copy.deepcopy(train_local_metrics["test_loss"]))

            # test data
            test_local_metrics = client.local_test(True)
            test_metrics["num_samples"].append(copy.deepcopy(test_local_metrics["test_total"]))
            test_metrics["num_correct"].append(copy.deepcopy(test_local_metrics["test_correct"]))
            test_metrics["losses"].append(copy.deepcopy(test_local_metrics["test_loss"]))

            # client_ws.append([round_idx,
            #                   client_idx,
            #                   train_metrics["losses"][client_idx] / train_metrics["num_samples"][client_idx],
            #                   train_metrics["num_correct"][client_idx] / train_metrics["num_samples"][client_idx],
            #                   time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())])

        # test on training dataset
        train_acc = sum(train_metrics["num_correct"]) / sum(train_metrics["num_samples"])
        train_loss = sum(train_metrics["losses"]) / sum(train_metrics["num_samples"])

        # test on test dataset
        test_acc = sum(test_metrics["num_correct"]) / sum(test_metrics["num_samples"])
        test_loss = sum(test_metrics["losses"]) / sum(test_metrics["num_samples"])

        stats = {"training_acc": train_acc, "training_loss": train_loss}

        # 绘制精度图
        accuracy_list.append(train_acc)
        loss_list.append(train_loss)
        plot_accuracy_and_loss(self, round_idx)

        if self.args.enable_wandb:
            wandb.log({"Train/Acc": train_acc, "round": round_idx})
            wandb.log({"Train/Loss": train_loss, "round": round_idx})

        mlops.log({"Train/Acc": train_acc, "round": round_idx})
        mlops.log({"Train/Loss": train_loss, "round": round_idx})
        logging.info(stats)

        stats = {"test_acc": test_acc, "test_loss": test_loss}
        if self.args.enable_wandb:
            wandb.log({"Test/Acc": test_acc, "round": round_idx})
            wandb.log({"Test/Loss": test_loss, "round": round_idx})

        mlops.log({"Test/Acc": test_acc, "round": round_idx})
        mlops.log({"Test/Loss": test_loss, "round": round_idx})
        logging.info(stats)

        return train_acc, train_loss, test_acc, test_loss

    def _local_test_on_validation_set(self, round_idx):

        logging.info("################local_test_on_validation_set : {}".format(round_idx))

        if self.val_global is None:
            self._generate_validation_set()

        client = self.client_list[0]
        client.update_local_dataset(0, None, self.val_global, None)
        # test data
        test_metrics = client.local_test(True)

        if self.args.dataset == "stackoverflow_nwp":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {"test_acc": test_acc, "test_loss": test_loss}
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})

        elif self.args.dataset == "stackoverflow_lr":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_pre = test_metrics["test_precision"] / test_metrics["test_total"]
            test_rec = test_metrics["test_recall"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {
                "test_acc": test_acc,
                "test_pre": test_pre,
                "test_rec": test_rec,
                "test_loss": test_loss,
            }
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Pre": test_pre, "round": round_idx})
                wandb.log({"Test/Rec": test_rec, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Pre": test_pre, "round": round_idx})
            mlops.log({"Test/Rec": test_rec, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})
        else:
            raise Exception("Unknown format to log metrics for dataset {}!" % self.args.dataset)

        logging.info(stats)


# 定义绘制精度图的函数
def plot_accuracy_and_loss(self, round_idx):
    plt.ion()

    print("accuracy_list: ", accuracy_list)
    print("loss_list: ", loss_list)

    plt.clf()
    plt.suptitle("FedGBTC" + "_[" + self.args.dataset + "]_NIID" + str(self.args.experiment_niid_level))

    # 第1个子图
    plt.subplot(2, 2, 1)
    plt.title("accuracy")
    plt.xlabel("num of epoch")
    plt.ylabel("value of accuracy")
    plt.plot(range(1, len(accuracy_list) + 1), accuracy_list, 'b-', linewidth=2)

    # 第2个子图
    plt.subplot(2, 2, 2)
    plt.title("loss")
    plt.xlabel("num of epoch")
    plt.ylabel("value of loss")
    plt.plot(range(1, len(loss_list) + 1), loss_list, 'b-', linewidth=2)

    # # 第3个子图
    # plt.subplot(2, 2, 3)
    # plt.title("k")
    # plt.xlabel("num of epoch")
    # plt.ylabel("value of k")
    # plt.plot(range(1, len(k_list) + 1), k_list, 'b-', linewidth=2)

    # # 第4个子图，使用条形图展示每个客户的被选择次数
    # plt.subplot(2, 2, 4)
    # plt.title("num of selected")
    # plt.xlabel("num of epoch")
    # plt.ylabel("value of num of selected")
    # plt.bar(range(1, len(client_selected_times) + 1), client_selected_times, width=0.5, fc='b')

    plt.tight_layout()
    plt.pause(0.005)
    plt.ioff()

    if (round_idx == self.args.comm_round - 1):
        plt.show()

    return
