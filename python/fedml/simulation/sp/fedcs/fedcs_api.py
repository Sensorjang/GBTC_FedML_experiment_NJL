import copy
import logging
import random
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import torch
import wandb
import openpyxl
import time
from fedml import mlops
from fedml.ml.trainer.trainer_creator import create_model_trainer
from .client import Client, client_upload_time, client_aggregation_time
import matplotlib.pyplot as plt

accuracy_list = []
loss_list = []
train_time_list = []
member_num_list = []
selfish_num_list = []

# 参数0
# 统计global_client_num_in_total个客户每个人的被选择次数
# client_selected_times = [0 for i in range(global_client_num_in_total)]
plt.figure(1, figsize=(20, 10))

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 删除默认创建的工作表（如果需要的话）
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])
# 创建一个工作表用于存放全局指标
global_metrics_ws = wb.create_sheet('Global Metrics')
# 写入全局指标的标头行
global_metrics_ws.append(['Round', 'Test Accuracy', 'Test Loss', 'Training Time', 'Member Num', 'Selfish Num'])

# 设置时间间隔（以秒为单位）
interval = 5


class AutoIncrementDict:
    def __init__(self):
        self._data = {}
        self._next_id = 0

    def add(self, value):
        self._data[self._next_id] = value
        current_id = self._next_id
        self._next_id += 1
        return current_id

    def remove(self, key):
        if key in self._data:
            del self._data[key]

    def get(self, key):
        return self._data.get(key, None)

    def __getitem__(self, key):
        return self.get(key)

    def __delitem__(self, key):
        self.remove(key)

    def __repr__(self):
        return repr(self._data)

    def items(self):
        return self._data.items()

    def keys(self):
        return self._data.keys()

    def values(self):
        return self._data.values()

    def remove_empty_values(self):
        keys_to_remove = [key for key, value in self._data.items() if not value]
        for key in keys_to_remove:
            del self._data[key]

    def ban_union(self, keys_to_remove):
        for key in keys_to_remove:
            del self._data[key]





class FedCSAPI(object):
    def __init__(self, args, device, dataset, model):
        self.device = device
        self.args = args

        [
            train_data_num,
            test_data_num,
            train_data_global,
            test_data_global,
            train_data_local_num_dict,
            train_data_local_dict,
            test_data_local_dict,
            class_num,
        ] = dataset
        self.client_num_in_total = args.client_num_in_total
        self.train_global = train_data_global
        self.test_global = test_data_global
        self.val_global = None
        self.train_data_num_in_total = train_data_num
        self.test_data_num_in_total = test_data_num
        self.client_list = []
        self.train_data_local_num_dict = train_data_local_num_dict
        self.train_data_local_dict = train_data_local_dict
        self.test_data_local_dict = test_data_local_dict

        # 历史联盟(客户历史所参与的,还没有盟主，字典存放联盟的id-客户对其偏好值)
        self.trust_threshold = args.trust_threshold
        self.member_range = tuple(args.member_range)
        self.trust_matrix = self.create_trust_graph()
        self.his_client_unions = {i: {} for i in range(self.client_num_in_total)}

        logging.info("model = {}".format(model))
        self.model_trainer = create_model_trainer(model, args)
        self.model = model
        logging.info("self.model_trainer = {}".format(self.model_trainer))

        self._setup_clients(
            train_data_local_num_dict, train_data_local_dict, test_data_local_dict, copy.deepcopy(self.model_trainer),
        )

    def _setup_clients(
        self, train_data_local_num_dict, train_data_local_dict, test_data_local_dict, model_trainer,
    ):
        logging.info("############setup_clients (START)#############")
        for client_idx in range(self.client_num_in_total):
            c = Client(
                client_idx,
                train_data_local_dict[client_idx],
                test_data_local_dict[client_idx],
                train_data_local_num_dict[client_idx],
                self.args,
                self.device,
                copy.deepcopy(model_trainer),
            )
            self.client_list.append(c)
        logging.info("############setup_clients (END)#############")
    def create_trust_graph(self, value_range=(0, 10), distrust_probability=0.01):
        """
        修改后的函数，以一定的概率生成表示完全不信任的负无穷值，并将对角线元素设置为0。
        :param value_range: 信任值的范围，为(min_value, max_value)
        :param distrust_probability: 完全不信任的概率
        :return: 带有表示不信任关系的随机信任图，对角线元素为0
        """
        shape = (self.client_num_in_total, self.client_num_in_total)
        min_value, max_value = value_range
        # 生成随机信任值数组
        trust_array = np.random.rand(*shape) * (max_value - min_value) + min_value
        # 以一定概率设置某些信任值为负无穷，表示完全不信任
        distrust_mask = np.random.rand(*shape) < distrust_probability
        trust_array[distrust_mask] = -np.inf
        # 将对角线元素设置为0
        np.fill_diagonal(trust_array, 0)
        return trust_array

    def cal_preference(self, i, union):  # 客户对联盟的偏好函数,暂时不考虑历史联盟 (计算时不考虑历史参与)
        pre = 0.0
        for j in union:
            if self.trust_matrix[i][j] == -np.inf:
                pre = -np.inf
                break
            # elif
            else:
                pre += self.trust_matrix[i][j]
        return pre

    def upgrade_union_uid(self, union, uid_old, uid_new, nid=None):  # 新旧联盟的交替(通常发生在某客户离开/加入某联盟，会产生新的联盟，此时需要修改其余成员的记录)
        for cid in union:  # 这里可选则将新增客户id传入,避免删除旧联盟id记录的操作
            if cid != nid:
                self.his_client_unions[cid].pop(uid_old)
            self.his_client_unions[cid][uid_new] = self.cal_preference(cid, union)

    def cal_trust_sum(self, union, cid):
        return np.sum([self.trust_matrix[i][cid] for i in union])

    def unions_formation(self):
        """
        生成初始联盟划分，同时避免联盟中存在信任值为负无穷的客户关系。
        """
        customer_ids = np.arange(self.client_num_in_total)
        np.random.shuffle(customer_ids)
        unions = AutoIncrementDict()
        client_selfish_list = []
        # 初始化联盟字典（自动计数id）
        for customer_id in customer_ids:
            added = False
            for uid, members in unions.items():
                can_add = True
                for member in members:
                    if (self.trust_matrix[customer_id, member] == -np.inf or
                            self.trust_matrix[member, customer_id] == -np.inf):
                        can_add = False
                        break
                if can_add:
                    unions[uid].append(customer_id)
                    self.his_client_unions[customer_id][uid] = None  # 此时联盟还未稳定，先不计算偏好值
                    added = True
                    break
            if not added:
                new_uid = unions.add([customer_id])
                self.his_client_unions[customer_id][new_uid] = None  # 此时联盟还未稳定，先不计算偏好值
        unions.remove_empty_values()  # 清除空联盟

        stable = False
        while not stable:
            stable = True  # 假设本轮次不再发生联盟变化
            for cid in range(self.client_num_in_total):  # 遍历每一位客户,初始将目前定义为最优
                uid_i = next(reversed(self.his_client_unions[cid].keys()))  # 找到当前客户i所在的联盟id
                best_pre_i = self.cal_preference(cid, unions[uid_i])  # 先计算客户对当前联盟的偏好值
                best_uid_i = uid_i
                for uid, union in unions.items():
                    if uid == best_uid_i:  # 避免重新评估当前联盟
                        continue
                    else:
                        pre = self.cal_preference(cid, union) \
                            if uid not in self.his_client_unions[cid] else 0  # 计算客户对该新联盟的偏好值（此时需要考虑历史联盟的问题）
                        if pre > best_pre_i:
                            best_pre_i = pre
                            best_uid_i = uid
                            stable = False  # 如果发生了联盟变化，则本轮次不再稳定

                if best_uid_i != uid_i:  # 如果找到了更好的联盟,双更新,也要更新原来两个联盟中客户绑定的uid
                    # 更新旧联盟
                    union_former = unions[uid_i]
                    union_former.remove(cid)
                    unions.remove(uid_i)
                    new_uid_former = unions.add(union_former)
                    self.upgrade_union_uid(union_former, uid_i, new_uid_former)
                    # 更新新联盟
                    union_latter = unions[best_uid_i]
                    unions.remove(best_uid_i)
                    union_latter.append(cid)
                    new_uid_latter = unions.add(union_latter)
                    self.his_client_unions[cid].pop(uid_i)
                    self.upgrade_union_uid(union_latter, best_uid_i, new_uid_latter, cid)
                unions.remove_empty_values()  # 清除空联盟

        # 新增代码：在最终联盟确定后，清除只有一个客户的联盟
        for uid, union in unions.items():  # 使用list来避免在遍历时修改字典
            for cid in union:  # 清除社会信任不足阈值的成员
                if self.cal_trust_sum(union, cid) < self.trust_threshold:
                    self.his_client_unions[cid].pop(uid)
                    client_selfish_list.append(cid)  # 因为社会信任不足被剔除
                    union.remove(cid)

        min_union = self.member_range[0]  # 不成盟，只用处理下界
        # 清除自私集群的联盟
        for uid, union in unions.items():
            if len(union) <= 3:
                for cid in union:
                    self.his_client_unions[cid].pop(uid)
                    client_selfish_list.append(cid)
        print(client_selfish_list)

        # 联盟人数恢复策略
        # min_uid = list(sorted(unions.items(), key=lambda x: len(x[1]), reverse=False))[0][0]  # 按照字典的值-联盟长度逆序排序
        # len_rest = len(client_selfish_list)
        # while len_rest > min_union:
        #     if len_rest - min_union >= min_union:
        #         new_union = []
        #         for i in range(min_union):
        #             new_union.append(client_selfish_list.pop())
        #         unions.add(new_union)  # 取出前5加入
        #         len_rest -= min_union
        #     else:
        #         cid = client_selfish_list.pop()
        #         unions[min_uid].append(cid)
        #         min_uid = list(sorted(unions.items(), key=lambda x: len(x[1]), reverse=False))[0][0]
        #         len_rest -= 1

        return client_selfish_list

    def train(self):
        logging.info("self.model_trainer = {}".format(self.model_trainer))
        w_global = self.model_trainer.get_model_params()
        mlops.log_training_status(mlops.ClientConstants.MSG_MLOPS_CLIENT_STATUS_TRAINING)
        mlops.log_aggregation_status(mlops.ServerConstants.MSG_MLOPS_SERVER_STATUS_RUNNING)
        mlops.log_round_info(self.args.comm_round, -1)
        # 这部分同GBTC，用于确定那些是自私客户
        mlops.event("信任图生成", event_started=True)
        print(self.trust_matrix)
        mlops.event("信任图生成", event_started=False)
        mlops.event("联盟生成", event_started=True)
        client_selfish_list = self.unions_formation()  # 初始联盟划分
        mlops.event("联盟生成", event_started=False)
        logging.info("client_banned_indexes = " + str(client_selfish_list))

        for round_idx in range(self.args.comm_round):
            logging.info("################Communication round : {}".format(round_idx))
            client_indexes = self._client_sampling(
                round_idx, self.args.client_num_in_total)

            logging.info("client_indexes = " + str(client_indexes))
            print(len(client_indexes))
            w_locals = []
            s_time = time.time()
            this_ban_cid = []
            for cid in client_indexes:
                client = self.client_list[cid]
                # 自私客户不训练
                if cid not in client_selfish_list:
                    type = '诚实'
                    mlops.event("train", event_started=True, event_value="轮次{}_客户id{}_类型{}".format(str(round_idx), str(cid), type))
                    w = client.train(copy.deepcopy(w_global))
                    mlops.event("train", event_started=False, event_value="轮次{}_客户id{}_类型{}".format(str(round_idx), str(cid), type))
                    w_locals.append((client.get_sample_number(), copy.deepcopy(w)))
                else:
                    type = '自私'
                    mlops.event("train", event_started=True, event_value="轮次{}_客户id{}_类型{}".format(str(round_idx), str(client.client_idx), type))
                    mlops.event("train", event_started=False, event_value="轮次{}_客户id{}_类型{}".format(str(round_idx), str(client.client_idx), type))
                    w_locals.append((client.get_sample_number(), copy.deepcopy(w_global)))
                    this_ban_cid.append(cid)
            e_time = time.time()
            train_time_list.append(e_time - s_time)  # 记录本轮的训练时间
            print(this_ban_cid)
            print(len(this_ban_cid))
            mlops.event("agg", event_started=True, event_value=str(round_idx))
            w_global = self._aggregate(w_locals)
            self.model_trainer.set_model_params(w_global)
            mlops.event("agg", event_started=False, event_value=str(round_idx))

            # test results
            # at last round
            train_acc, train_loss, test_acc, test_loss = 0, 0, 0, 0
            if round_idx == self.args.comm_round - 1:
                self._local_test_on_all_clients(round_idx)
            # per {frequency_of_the_test} round
            elif round_idx % self.args.frequency_of_the_test == 0:
                if self.args.dataset.startswith("stackoverflow"):
                    self._local_test_on_validation_set(round_idx)
                else:
                    self._local_test_on_all_clients(round_idx)

            mlops.log_round_info(self.args.comm_round, round_idx)
            selfish_num = len(client_selfish_list)
            selfish_num_list.append(selfish_num)
            member_num_list.append(self.client_num_in_total - selfish_num)
            # round_ws.append([round_idx,
            #                     train_loss,
            #                     train_acc,
            #                     time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),
            #                     str(client_indexes),
            #                     str(client_selected_times)])

        # 填充数据到工作表
        for i in range(len(accuracy_list)):
            # 轮次号，测试精度，测试损失，训练时间
            round_num = i + 1  # 轮次号从1开始
            accuracy = accuracy_list[i]
            loss = loss_list[i]
            train_time = train_time_list[i]
            member_num = member_num_list[i]
            selfish_num = selfish_num_list[i]
            global_metrics_ws.append([round_num, accuracy, loss, train_time, member_num, selfish_num])

        # 保存Excel文件到self.args.excel_save_path+文件名
        wb.save(
            self.args.excel_save_path + self.args.model + "_[" + self.args.dataset + "]_FedCS_training_results_NIID" + str(
                self.args.experiment_niid_level) + ".xlsx")
        # 休眠一段时间，以便下一个循环开始前有一些时间
        # time.sleep(interval)

        mlops.log_training_finished_status()
        mlops.log_aggregation_finished_status()

    def _client_sampling(self, round_idx, client_num_in_total):
        # Calculate the estimated time elapsed for each client
        Tcs = 1  # Client Selection step time
        Td = 0  # Total time for clients in S
        Theta = 0  # Current time
        Tround = 10  # Maximum time for each round
        K0 = list(range(client_num_in_total))  # Set of randomly selected clients
        client_indexes = []  # Selected clients
        t = Tcs  # Current time
        while len(K0) > 0 and t < Tround:
            x = None
            max_val = float('-inf')
            for k in K0:
                Td_S_k = Td + self.client_list[k].estimate_upload_time(round_idx)  # Estimated time for k to upload its model
                Theta_S_k = Theta + Td_S_k  # Estimated time for k to complete the whole process
                val = Tcs + Td_S_k + max(0, Theta_S_k - Theta)  # Estimated time for k to join S
                if val > max_val:
                    max_val = val
                    x = k
            if x is None:
                break
            Td += self.client_list[x].estimate_upload_time(round_idx)
            Theta0 = Theta + self.client_list[x].estimate_upload_time(round_idx)
            t = Tcs + Td + max(0, Theta0 - Theta) + self.client_list[x].estimate_aggregation_time(round_idx)
            if t < Tround:
                client_indexes.append(x)
                K0.remove(x)
                Theta = Theta0
        logging.info("Selected clients: %s" % str(client_indexes))
        return client_indexes

    def _generate_validation_set(self, num_samples=10000):
        test_data_num = len(self.test_global.dataset)
        sample_indices = random.sample(range(test_data_num), min(num_samples, test_data_num))
        subset = torch.utils.data.Subset(self.test_global.dataset, sample_indices)
        sample_testset = torch.utils.data.DataLoader(subset, batch_size=self.args.batch_size)
        self.val_global = sample_testset

    def _aggregate(self, w_locals):
        training_num = 0
        for idx in range(len(w_locals)):
            (sample_num, averaged_params) = w_locals[idx]
            training_num += sample_num

        (sample_num, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            print("______k = " + str(k))
            for i in range(0, len(w_locals)):
                local_sample_number, local_model_params = w_locals[i]
                w = local_sample_number / training_num
                if i == 0:
                    averaged_params[k] = local_model_params[k] * w
                else:
                    averaged_params[k] += local_model_params[k] * w
        return averaged_params

    def _aggregate_noniid_avg(self, w_locals):
        """
        The old aggregate method will impact the model performance when it comes to Non-IID setting
        Args:
            w_locals:
        Returns:
        """
        (_, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            temp_w = []
            for (_, local_w) in w_locals:
                temp_w.append(local_w[k])
            averaged_params[k] = sum(temp_w) / len(temp_w)
        return averaged_params

    def local_test_thread(self, cid):
        return self.client_list[cid].local_test(True)
    def _local_test_on_all_clients(self, round_idx):

        logging.info("################local_test_on_all_clients : {}".format(round_idx))

        test_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        with ThreadPoolExecutor(max_workers=self.client_num_in_total) as executor:  # 多线程全客户测试（修复之前错误的测试逻辑，仅在0客户上测试）
            futures = []
            for cid in range(self.client_num_in_total):
                future = executor.submit(self.local_test_thread, cid)
                futures.append(future)
            for future in futures:
                test_metrics_cid = future.result()
                test_metrics["num_samples"].append(test_metrics_cid["test_total"])
                test_metrics["num_correct"].append(test_metrics_cid["test_correct"])
                test_metrics["losses"].append(test_metrics_cid["test_loss"])

        # test on test dataset
        test_acc = sum(test_metrics["num_correct"]) / sum(test_metrics["num_samples"])
        test_loss = sum(test_metrics["losses"]) / sum(test_metrics["num_samples"])

        stats = {"test_acc": test_acc, "test_loss": test_loss}
        if self.args.enable_wandb:
            wandb.log({"Test/Acc": test_acc, "round": round_idx})
            wandb.log({"Test/Loss": test_loss, "round": round_idx})

        mlops.log({"Test/Acc": test_acc, "round": round_idx})
        mlops.log({"Test/Loss": test_loss, "round": round_idx})
        logging.info(stats)
        # 绘制精度图(存入测试图)
        accuracy_list.append(test_acc)
        loss_list.append(test_loss)
        plot_accuracy_and_loss(self, round_idx)

        return test_acc, test_loss

    def _local_test_on_validation_set(self, round_idx):

        logging.info("################local_test_on_validation_set : {}".format(round_idx))

        if self.val_global is None:
            self._generate_validation_set()

        client = self.client_list[0]
        client.update_local_dataset(0, None, self.val_global, None)
        # test data
        test_metrics = client.local_test(True)

        if self.args.dataset == "stackoverflow_nwp":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {"test_acc": test_acc, "test_loss": test_loss}
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})

        elif self.args.dataset == "stackoverflow_lr":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_pre = test_metrics["test_precision"] / test_metrics["test_total"]
            test_rec = test_metrics["test_recall"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {
                "test_acc": test_acc,
                "test_pre": test_pre,
                "test_rec": test_rec,
                "test_loss": test_loss,
            }
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Pre": test_pre, "round": round_idx})
                wandb.log({"Test/Rec": test_rec, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Pre": test_pre, "round": round_idx})
            mlops.log({"Test/Rec": test_rec, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})
        else:
            raise Exception("Unknown format to log metrics for dataset {}!" % self.args.dataset)

        logging.info(stats)

# 定义绘制精度图的函数
def plot_accuracy_and_loss(self, round_idx):
    plt.ion()

    print("accuracy_list: ", accuracy_list)
    print("loss_list: ", loss_list)

    plt.clf()
    plt.suptitle("FedCS" + "_[" + self.args.dataset +"]_NIID"+ str(self.args.experiment_niid_level))

    # 第1个子图
    plt.subplot(2, 3, 1)
    plt.title("accuracy")
    plt.xlabel("num of epoch")
    plt.ylabel("value of accuracy")
    plt.plot(range(1, len(accuracy_list)+1), accuracy_list, 'b-', linewidth=2)

    # 第2个子图
    plt.subplot(2, 3, 2)
    plt.title("loss")
    plt.xlabel("num of epoch")
    plt.ylabel("value of loss")
    plt.plot(range(1, len(loss_list)+1), loss_list, 'b-', linewidth=2)

    # # 第3个子图，使用条形图展示每个客户的被选择次数
    # plt.subplot(2, 3, 3)
    # plt.title("num of selected")
    # plt.xlabel("num of epoch")
    # plt.ylabel("value of num of selected")
    # plt.bar(range(1, len(client_selected_times)+1), client_selected_times, width=0.5, fc='b')

    # # 第4个子图，使用折线图展示每个客户的预估上传时间
    # plt.subplot(2, 3, 4)
    # plt.title("estimated upload time")
    # plt.xlabel("Clients' Indexes")
    # plt.ylabel("value of estimated upload time")
    # for i in range(len(client_upload_time)):
    #     plt.plot(range(1, global_client_num_in_total+1), client_upload_time[i], label="Round {}".format(i + 1))
    # plt.legend()

    # # 第5个子图，使用折线图展示每个客户的预估聚合时间
    # plt.subplot(2, 3, 5)
    # plt.title("estimated aggregation time")
    # plt.xlabel("Clients' Indexes")
    # plt.ylabel("value of estimated aggregation time")
    # for i in range(len(client_aggregation_time)):
    #     plt.plot(range(1, global_client_num_in_total+1), client_aggregation_time[i], label="Round {}".format(i + 1))
    # plt.legend()

    plt.tight_layout()
    plt.pause(0.03)
    plt.ioff()

    if (round_idx == self.args.comm_round - 1):
        plt.show()

    return