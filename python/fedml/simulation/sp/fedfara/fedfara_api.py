import copy
import logging
import math
import random
from collections import OrderedDict

import matplotlib.pyplot as plt
import torch
import wandb
from fedml import mlops
from fedml.ml.trainer.trainer_creator import create_model_trainer

from .client import Client
import openpyxl
import time

global_client_num_in_total = 60
global_client_num_per_round = 30

accuracy_list = []
loss_list = []
k_list = []
# 参数0
# 统计global_client_num_in_total个客户每个人的被选择次数
client_selected_times = [0 for i in range(global_client_num_in_total)]
plt.figure(1, figsize=(16, 4))

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 创建工作表
client_ws = wb.create_sheet('Clients Info')
# 写入损失指标的标头行
client_ws.append(['Round', 'ClientIdx', 'Loss', 'Accuracy', 'Time'])
# 创建工作表
round_ws = wb.create_sheet('Round Info')
# 写入精度指标的标头行
round_ws.append(['Round', 'Loss', 'Accuracy', 'Time', 'Selected Client Indexs', 'Total Selected Client Times', 'K'])
# 创建工作表
bid_quality_ws = wb.create_sheet('Bid and Quality Info')
# 写入精度指标的标头行
bid_quality_ws.append(['Round', 'ClientIdx', 'Bid', 'Quality Score'])
# 设置时间间隔（以秒为单位）
interval = 5

class FedFARAAPI(object):
    def __init__(self, args, device, dataset, model):
        self.device = device
        self.args = args
        [
            train_data_num,
            test_data_num,
            train_data_global,
            test_data_global,
            train_data_local_num_dict,
            train_data_local_dict,
            test_data_local_dict,
            class_num,
        ] = dataset

        self.train_global = train_data_global
        self.test_global = test_data_global
        self.val_global = None
        self.train_data_num_in_total = train_data_num
        self.test_data_num_in_total = test_data_num
        self.class_num = class_num
        self.args.client_num_in_total = global_client_num_in_total # added
        self.client_list = []
        self.banned_client_indexs = []
        self.client_quility_list = []
        self.train_data_local_num_dict = train_data_local_num_dict
        self.train_data_local_dict = train_data_local_dict
        self.test_data_local_dict = test_data_local_dict

        logging.info("model = {}".format(model))
        self.model_trainer = create_model_trainer(model, args)
        self.model = model
        logging.info("self.model_trainer = {}".format(self.model_trainer))

        self._setup_clients(
            train_data_local_num_dict, train_data_local_dict, test_data_local_dict, self.model_trainer,
        )

    def _setup_clients(
            self, train_data_local_num_dict, train_data_local_dict, test_data_local_dict, model_trainer,
    ):
        logging.info("############setup_clients (START)#############")
        # 参数1
        for client_idx in range(global_client_num_in_total):
            c = Client(
                client_idx,
                train_data_local_dict[client_idx],
                test_data_local_dict[client_idx],
                train_data_local_num_dict[client_idx],
                self.args,
                self.device,
                model_trainer,
            )
            self.client_list.append(c)
        logging.info("############setup_clients (END)#############")

    def train(self):
        logging.info("self.model_trainer = {}".format(self.model_trainer))
        w_global = self.model_trainer.get_model_params()
        mlops.log_training_status(mlops.ClientConstants.MSG_MLOPS_CLIENT_STATUS_TRAINING)
        mlops.log_aggregation_status(mlops.ServerConstants.MSG_MLOPS_SERVER_STATUS_RUNNING)
        mlops.log_round_info(self.args.comm_round, -1)

        for round_idx in range(self.args.comm_round):

            # 定义局部变量client_list,其元素是从全局的client_list中剔除ban的client
            client_list = []
            for client in self.client_list:
                if client.client_idx not in self.banned_client_indexs:
                    client_list.append(client)

            logging.info("################Communication round : {}".format(round_idx))

            w_locals = []

            """
            for scalability: following the original FedAvg algorithm, we uniformly sample a fraction of clients in each round.
            Instead of changing the 'Client' instances, our implementation keeps the 'Client' instances and then updates their local dataset 
            """
            client_indexes = self._client_sampling(  # round 0 select all, then select client which is not banned
                round_idx, client_list, self.banned_client_indexs
            )

            logging.info("client_indexes = " + str(client_indexes))

            for idx, client in enumerate(client_list):
                # update dataset
                # 如果idx是client_indexes的一部分，那么就更新这个client的数据集
                if client.client_idx in client_indexes:
                    client_idx = client.client_idx
                    client.update_local_dataset(
                        client_idx,
                        self.train_data_local_dict[client_idx],
                        self.test_data_local_dict[client_idx],
                        self.train_data_local_num_dict[client_idx],
                    )

                    # train on new dataset
                    mlops.event("train", event_started=True,
                                event_value="{}_{}".format(str(round_idx), str(client.client_idx)))
                    w = client.train(copy.deepcopy(w_global))
                    mlops.event("train", event_started=False,
                                event_value="{}_{}".format(str(round_idx), str(client.client_idx)))
                    # self.logging.info("local weights = " + str(w))
                    w_locals.append((client.get_sample_number(), copy.deepcopy(w), client.client_idx))

            # 定义 similarity_scores 是一个map，存储每个client的index与w_global的余弦相似度
            similarity_scores = []
            # 遍历w_locals中的所有模型，计算它们与w_global的余弦相似度作为距离
            # 由于w_locals中的模型是按照client的顺序排列的，因此可以直接使用client的index来获取对应的模型
            # print("A\n")
            for i in range(len(w_locals)):
                # print("B\n")
                # 将张量展平为一维
                if self.args.model == "cnn":
                    local_flatten = torch.flatten(w_locals[i][1]['conv2d_1.weight'])
                    global_flatten = torch.flatten(w_global['conv2d_1.weight'])
                elif self.args.model == "rnn":
                    local_flatten = torch.flatten(w_locals[i][1]['embeddings.weight'])
                    global_flatten = torch.flatten(w_global['embeddings.weight'])
                elif self.args.model == "resnet18_gn":
                    local_flatten = torch.flatten(w_locals[i][1]['conv1.weight'])
                    global_flatten = torch.flatten(w_global['conv1.weight'])
                elif self.args.model == "cnn_web":
                    local_flatten = torch.flatten(w_locals[i][1]['conv1.weight'])
                    global_flatten = torch.flatten(w_global['conv1.weight'])
                else:
                    local_flatten = torch.flatten(w_locals[i][1]['conv1.weight'])
                    global_flatten = torch.flatten(w_global['conv1.weight'])

                # 计算点积
                dot_product = torch.dot(local_flatten, global_flatten)

                # 计算范数
                local_norm = torch.norm(local_flatten)
                global_norm = torch.norm(global_flatten)

                # 计算余弦相似度
                cosine_similarity = dot_product / (local_norm * global_norm)
                # print("C\n")
                cos = torch.nn.CosineSimilarity(dim=0, eps=1e-6)
                # 打印w_locals[i][1]、w_global
                # print("w_locals[i][1] = " + str(w_locals[i][1]))
                # print("w_global = " + str(w_global))
                distance = cosine_similarity.item()
                # print("D\n")
                similarity_scores.append(abs(distance))

            # print("E\n")
            # 使用阈值筛选出低于阈值的一部分，将它们对应的client加入ban列表
            if round_idx == 0:
                # print("F\n")
                # 借助采用经典莱茵达准则过滤异常值的方法，将距离中的异常值剔除(加入ban列表)
                mean_similarity_scores = sum(similarity_scores) / len(similarity_scores)
                variance = sum([(x - mean_similarity_scores) ** 2 for x in similarity_scores]) / (
                        len(similarity_scores) - 1)
                sigma = math.sqrt(variance)
                threshold = 3 * sigma
                for i in range(len(similarity_scores)):
                    print("similarity_scores[i] = " + str(abs(similarity_scores[i])))
                    print("threshold = " + str(threshold))
                    # if self.args.dataset in ["fed_shakespeare"]:
                    #     if abs(similarity_scores[i]) > threshold:
                    #         self.banned_client_indexs.append(i)
                    # elif abs(similarity_scores[i]) <= threshold:
                    if abs(similarity_scores[i]) <= threshold:
                        self.banned_client_indexs.append(i)
            # print("G\n")
            print("ban_list = " + str(self.banned_client_indexs))

            # Step 1: 数据归一化和非负化处理
            normalized_scores = []
            # 打印min(similarity_scores)) max(similarity_scores) similarity_scores
            # print("min(similarity_scores)) = " + str(min(similarity_scores)))
            # print("max(similarity_scores) = " + str(max(similarity_scores)))
            print("similarity_scores = " + str(similarity_scores))
            for score in similarity_scores:
                max_min_diff = max(similarity_scores) - min(similarity_scores)
                if max_min_diff == 0:
                    normalized_score = 1.0  # 分母为0时，将归一化分数设置为1.0
                else:
                    normalized_score = 1 + (score - min(similarity_scores)) / max_min_diff
                normalized_scores.append(normalized_score)
            # print("H\n")
            print("normalized_scores = " + str(normalized_scores))
            # Step 2: 计算每个客户在相似度指标上的权重
            total_normalized_score = sum(normalized_scores)
            eweights = [(normalized_score / total_normalized_score) for normalized_score in normalized_scores]
            # print("I\n")
            # Step 3: 计算每个客户的熵权值
            # 打印eweights、len(similarity_scores)
            print("eweights = " + str(eweights))
            num_clients = int(len(similarity_scores))
            print("len(similarity_scores) = " + str(num_clients))
            # 数据集中数据类别（class）总数
            # 定义二维数组x[i,j]，分别是客户端index i、数据类别j、客户i数据中类别为j的占比（概率）
#参数
            x_list = [[] for _ in range(global_client_num_in_total)]
            # 遍历client_indexes
            for i in client_indexes:
                for j in range(self.class_num):

                    this_client = None
                    this_sample_number = 0

                    for idx, client in enumerate(client_list):
                        # update dataset
                        # 如果idx是client_indexes的一部分，那么就更新这个client的数据集
                        if client.client_idx == i:
                            this_client = client

                    for (sample_number, _, client_idx) in w_locals:
                        if client_idx == i:
                            this_sample_number = sample_number

                    x_list[i].append(
                        (this_client.get_sample_class_number(j) + 1) / (this_sample_number + self.class_num))
            # 打印x_list
            # print("x_list = " + str(x_list))

            entropy = []
            for i in client_indexes:
                sum_entropy = 0
                for j in range(self.class_num):
                    sum_entropy -= x_list[i][j] * math.log(x_list[i][j])
                entropy.append(sum_entropy)
            sum_entropy = sum([(1 - entropy) for entropy in entropy])
            wi = [((1 - entropy) / sum_entropy) for entropy in entropy]
            # print("J\n")
            # Step 4: 计算每个客户的数据质量得分
            quality_scores = [0 for _ in range(global_client_num_in_total)]
            for i in range(len(w_locals)):
                quality_score = wi[i] * w_locals[i][0]
                # quality_scores.append(quality_score)
                quality_scores[w_locals[i][2]] = quality_score
            # print("K\n")

            # 参数2
            # 定义总的预算上限
            total_budget = 3300
            # 定义每个client的报价列表bid_list,然后初始化它们的报价随机的10~99之间的整数
            # 参数3
            bid_list = [0 for _ in range(global_client_num_in_total)]
            # for i in range(len(w_locals)):
            for i in range(global_client_num_in_total):
                bid = random.randint(10, 99)
                # bid_list.append(bid)
                bid_list[i] = bid
            print("total_budget = " + str(total_budget))
            print("bid_list = " + str(bid_list))

            # print("L\n")
            # 定义一个三元组list，存储client index、报价、数据质量得分
            client_info = []
            for i in range(len(w_locals)):
                client_info.append((client_indexes[i], bid_list[client_indexes[i]], quality_scores[client_indexes[i]]))
                bid_quality_ws.append([round_idx, client_indexes[i], bid_list[client_indexes[i]], quality_scores[client_indexes[i]]])
            # 输出client_info
            logging.info("client_info = " + str(client_info))
            # 按照数据质量得分/报价的比值从大到小排序
            client_info.sort(key=lambda x: x[2] / x[1], reverse=True)
            # 输出排序后的client_info
            logging.info("client_info(sorted) = " + str(client_info))

            # 从大到小遍历client_info，直到达到预算上限
            # print("M\n")
            k = 0
            # 按顺序遍历client_info的每一个元组
            sum_quality = 0
            for i, (idx, bid, quality_score) in enumerate(client_info):
                sum_quality += quality_score
                # print("clients += " + str(idx))
                # print("sum_quality = " + str(sum_quality))
                now_budget = sum_quality * bid_list[idx] / quality_scores[idx]
                # print("bid = " + str(bid_list[i]))
                print("now_budget = " + str(now_budget))
                k = i
                if now_budget > total_budget:
                    break
            print("###############[k] = " + str(k))
            k_list.append(k)

            # print("N\n")
            # 将k到len(w_locals)的client从client_indexes和w_locals中剔除
            # for i 从len(w_locals)到k
            remove_index_record = []
            for i in range(len(w_locals) - 1, k, -1):
                # 拒绝第index=client_info[i][0]个client参与聚合
                # print("remove element i : " + str(client_info[i][0]))
                # 取出client_info[i]的三元组的第1个元素，即client的index
                client_indexes.remove(client_info[i][0])
                remove_index_record.append(client_info[i][0])
            # sort remove_index_record from big to small
            remove_index_record.sort(reverse=True)
            # 顺序遍历remove_index_record每个元素
            for i in remove_index_record:
                # 移除w_locals第idx个元素
                # print("w_locals" + str(w_locals))
                for idx, (sample_number,w,client_idx) in enumerate(w_locals):
                    if client_idx == i:
                        eweights.pop(idx)
                        w_locals.pop(idx)
            # GBTC_final_indexes展示的是最终被选中的client的index
            logging.info("GBTC_final_indexes = " + str(client_indexes))
            # 借助client_selected_times统计global_client_num_in_total个客户每个人的被选择次数
            for i in client_indexes:
                client_selected_times[i] += 1

            # print("O\n")
            # update global weights
            mlops.event("agg", event_started=True, event_value=str(round_idx))
            w_global = self._aggregate(w_locals, eweights)

            self.model_trainer.set_model_params(w_global)
            mlops.event("agg", event_started=False, event_value=str(round_idx))

            # test results
            # at last round
            train_acc, train_loss, test_acc, test_loss = 0, 0, 0, 0
            if round_idx == self.args.comm_round - 1:
                train_acc, train_loss, test_acc, test_loss = self._local_test_on_all_clients(round_idx)
            # per {frequency_of_the_test} round
            elif round_idx % self.args.frequency_of_the_test == 0:
                if self.args.dataset.startswith("stackoverflow"):
                    self._local_test_on_validation_set(round_idx)
                else:
                    train_acc, train_loss, test_acc, test_loss = self._local_test_on_all_clients(round_idx)

            mlops.log_round_info(self.args.comm_round, round_idx)

            round_ws.append([round_idx,
                                train_loss,
                                train_acc,
                                time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),
                                str(client_indexes),
                                str(client_selected_times),
                                k])

            # 保存Excel文件到self.args.excel_save_path+文件名
            wb.save(self.args.excel_save_path + self.args.model + "_[" + self.args.dataset +"]_GBTC_training_results_NIID"+ str(self.args.experiment_niid_level) +".xlsx")
            # 休眠一段时间，以便下一个循环开始前有一些时间
            time.sleep(interval)

        mlops.log_training_finished_status()
        mlops.log_aggregation_finished_status()

    def _client_sampling(self, round_idx, client_list, banned_client_indexs):
        # round 0 select all, then select client which is not banned
        if round_idx == 0:
            client_indexes = [client.client_idx for client in client_list]
        else:
            client_indexes = [client.client_idx for client in client_list if
                              client.client_idx not in banned_client_indexs]

        # if client_num_in_total == client_num_per_round:
        #     client_indexes = [client_index for client_index in range(client_num_in_total)]
        # else:
        #     num_clients = min(client_num_per_round, client_num_in_total)
        #     np.random.seed(round_idx)  # make sure for each comparison, we are selecting the same clients each round
        #     client_indexes = np.random.choice(range(client_num_in_total), num_clients, replace=False)
        # logging.info("client_indexes = %s" % str(client_indexes))
        return client_indexes

    def _generate_validation_set(self, num_samples=10000):
        test_data_num = len(self.test_global.dataset)
        sample_indices = random.sample(range(test_data_num), min(num_samples, test_data_num))
        subset = torch.utils.data.Subset(self.test_global.dataset, sample_indices)
        sample_testset = torch.utils.data.DataLoader(subset, batch_size=self.args.batch_size)
        self.val_global = sample_testset

    def _aggregate(self, w_locals, eweights):
        # training_num = 0
        # for idx in range(len(w_locals)):
        #     (sample_num, averaged_params) = w_locals[idx]
        #     training_num += sample_num
        print("_aggregate.w_locals.length = " + str(len(w_locals)) + "\n" + "_aggregate.eweights.length = " + str(
            len(eweights)))

        eweights_sum = sum(eweights)

        (sample_num, averaged_params, _) = w_locals[0]
        for k in averaged_params.keys():
            print("______k = " + str(k))
            for i in range(0, len(w_locals)):
                # local_sample_number, local_model_params = w_locals[i]
                local_model_params = w_locals[i][1]
                local_wight_number = eweights[i]

                w = local_wight_number / eweights_sum

                # print("w = " + str(w))

                if i == 0:
                    averaged_params[k] = local_model_params[k] * w
                else:
                    averaged_params[k] += local_model_params[k] * w
        return averaged_params

    def _aggregate_resnet(self, w_locals, eweights): # 弃用
        eweights_sum = sum(eweights)

        averaged_params = OrderedDict()
        for (_, local_model_params), w in zip(w_locals, eweights):
            for key, param in local_model_params.items():
                if key not in averaged_params:
                    averaged_params[key] = []

                    # 获取对应权重
                local_wight = w / eweights_sum

                # 根据类型分别计算加权平均
                if 'conv' in key:
                    averaged_params[key] += local_wight * param.clone().detach()
                elif 'bn' in key:
                    averaged_params[key] += local_wight * param.clone().detach()
                elif 'fc' in key:
                    averaged_params[key] += local_wight * param.clone().detach()

        for key in averaged_params:
            averaged_params[key] /= eweights_sum

        return averaged_params

    def _aggregate_noniid_avg(self, w_locals):
        """
        The old aggregate method will impact the model performance when it comes to Non-IID setting
        Args:
            w_locals:
        Returns:
        """
        (_, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            temp_w = []
            for (_, local_w) in w_locals:
                temp_w.append(local_w[k])
            averaged_params[k] = sum(temp_w) / len(temp_w)
        return averaged_params

    def _local_test_on_all_clients(self, round_idx):

        logging.info("################local_test_on_all_clients : {}".format(round_idx))

        train_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        test_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        client = self.client_list[0]

        for client_idx in range(self.args.client_num_in_total):
            """
            Note: for datasets like "fed_CIFAR100" and "fed_shakespheare",
            the training client number is larger than the testing client number
            """
            if self.test_data_local_dict[client_idx] is None:
                continue
            client.update_local_dataset(
                0,
                self.train_data_local_dict[client_idx],
                self.test_data_local_dict[client_idx],
                self.train_data_local_num_dict[client_idx],
            )
            # train data
            train_local_metrics = client.local_test(False)
            train_metrics["num_samples"].append(copy.deepcopy(train_local_metrics["test_total"]))
            train_metrics["num_correct"].append(copy.deepcopy(train_local_metrics["test_correct"]))
            train_metrics["losses"].append(copy.deepcopy(train_local_metrics["test_loss"]))

            # test data
            test_local_metrics = client.local_test(True)
            test_metrics["num_samples"].append(copy.deepcopy(test_local_metrics["test_total"]))
            test_metrics["num_correct"].append(copy.deepcopy(test_local_metrics["test_correct"]))
            test_metrics["losses"].append(copy.deepcopy(test_local_metrics["test_loss"]))


            client_ws.append([round_idx,
                              client_idx,
                              train_metrics["losses"][client_idx] / train_metrics["num_samples"][client_idx],
                              train_metrics["num_correct"][client_idx] / train_metrics["num_samples"][client_idx],
                              time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())])

        # test on training dataset
        train_acc = sum(train_metrics["num_correct"]) / sum(train_metrics["num_samples"])
        train_loss = sum(train_metrics["losses"]) / sum(train_metrics["num_samples"])

        # test on test dataset
        test_acc = sum(test_metrics["num_correct"]) / sum(test_metrics["num_samples"])
        test_loss = sum(test_metrics["losses"]) / sum(test_metrics["num_samples"])

        stats = {"training_acc": train_acc, "training_loss": train_loss}

        # 绘制精度图
        accuracy_list.append(train_acc)
        loss_list.append(train_loss)
        plot_accuracy_and_loss(self, round_idx)

        if self.args.enable_wandb:
            wandb.log({"Train/Acc": train_acc, "round": round_idx})
            wandb.log({"Train/Loss": train_loss, "round": round_idx})

        mlops.log({"Train/Acc": train_acc, "round": round_idx})
        mlops.log({"Train/Loss": train_loss, "round": round_idx})
        logging.info(stats)

        stats = {"test_acc": test_acc, "test_loss": test_loss}
        if self.args.enable_wandb:
            wandb.log({"Test/Acc": test_acc, "round": round_idx})
            wandb.log({"Test/Loss": test_loss, "round": round_idx})

        mlops.log({"Test/Acc": test_acc, "round": round_idx})
        mlops.log({"Test/Loss": test_loss, "round": round_idx})
        logging.info(stats)

        return train_acc, train_loss, test_acc, test_loss

    def _local_test_on_validation_set(self, round_idx):

        logging.info("################local_test_on_validation_set : {}".format(round_idx))

        if self.val_global is None:
            self._generate_validation_set()

        client = self.client_list[0]
        client.update_local_dataset(0, None, self.val_global, None)
        # test data
        test_metrics = client.local_test(True)

        if self.args.dataset == "stackoverflow_nwp":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {"test_acc": test_acc, "test_loss": test_loss}
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})

        elif self.args.dataset == "stackoverflow_lr":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_pre = test_metrics["test_precision"] / test_metrics["test_total"]
            test_rec = test_metrics["test_recall"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {
                "test_acc": test_acc,
                "test_pre": test_pre,
                "test_rec": test_rec,
                "test_loss": test_loss,
            }
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Pre": test_pre, "round": round_idx})
                wandb.log({"Test/Rec": test_rec, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Pre": test_pre, "round": round_idx})
            mlops.log({"Test/Rec": test_rec, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})
        else:
            raise Exception("Unknown format to log metrics for dataset {}!" % self.args.dataset)

        logging.info(stats)


# 定义绘制精度图的函数
def plot_accuracy_and_loss(self, round_idx):
    plt.ion()

    print("accuracy_list: ", accuracy_list)
    print("loss_list: ", loss_list)

    plt.clf()
    plt.suptitle("FedGBTC" + "_[" + self.args.dataset +"]_NIID"+ str(self.args.experiment_niid_level))

    # 第1个子图
    plt.subplot(2, 2, 1)
    plt.title("accuracy")
    plt.xlabel("num of epoch")
    plt.ylabel("value of accuracy")
    plt.plot(range(1, len(accuracy_list) + 1), accuracy_list, 'b-', linewidth=2)

    # 第2个子图
    plt.subplot(2, 2, 2)
    plt.title("loss")
    plt.xlabel("num of epoch")
    plt.ylabel("value of loss")
    plt.plot(range(1, len(loss_list) + 1), loss_list, 'b-', linewidth=2)

    # 第3个子图
    plt.subplot(2, 2, 3)
    plt.title("k")
    plt.xlabel("num of epoch")
    plt.ylabel("value of k")
    plt.plot(range(1, len(k_list) + 1), k_list, 'b-', linewidth=2)

    # 第4个子图，使用条形图展示每个客户的被选择次数
    plt.subplot(2, 2, 4)
    plt.title("num of selected")
    plt.xlabel("num of epoch")
    plt.ylabel("value of num of selected")
    plt.bar(range(1, len(client_selected_times) + 1), client_selected_times, width=0.5, fc='b')

    plt.tight_layout()
    plt.pause(0.005)
    plt.ioff()

    if (round_idx == self.args.comm_round - 1):
        plt.show()

    return
